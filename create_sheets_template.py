#!/usr/bin/env python3
"""
DW24 Empfehlungsprogramm – Google Sheets Template erstellen
Erstellt eine XLSX-Datei mit 4 Blättern: Partner, Empfehlungen, Auszahlungen, Dashboard
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ===== Farben & Styles =====
DW24_DUNKEL = "1A1A2E"
DW24_ORANGE = "FF8C00"
DW24_WEISS = "FFFFFF"
GRUEN = "22C55E"
GELB = "F59E0B"
ROT = "EF4444"
BLAU = "3B82F6"
ORANGE_HELL = "FB923C"
GRAU_BG = "F1F5F9"

# Header-Style
header_font = Font(name="Calibri", bold=True, color=DW24_WEISS, size=11)
header_fill = PatternFill(start_color=DW24_DUNKEL, end_color=DW24_DUNKEL, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Bedingte Formatierung: Hintergrundfarben
fill_gruen = PatternFill(start_color=GRUEN, end_color=GRUEN, fill_type="solid")
fill_gelb = PatternFill(start_color=GELB, end_color=GELB, fill_type="solid")
fill_rot = PatternFill(start_color=ROT, end_color=ROT, fill_type="solid")
fill_blau = PatternFill(start_color=BLAU, end_color=BLAU, fill_type="solid")
fill_orange = PatternFill(start_color=DW24_ORANGE, end_color=DW24_ORANGE, fill_type="solid")
fill_orange_hell = PatternFill(start_color=ORANGE_HELL, end_color=ORANGE_HELL, fill_type="solid")

font_weiss = Font(color=DW24_WEISS, bold=True, size=10)


def style_header_row(ws, num_cols):
    """Header-Zeile formatieren"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    # Zeile fixieren
    ws.freeze_panes = "A2"


def auto_column_width(ws, headers):
    """Spaltenbreite automatisch anpassen"""
    for i, header in enumerate(headers, 1):
        # Mindestbreite basierend auf Header-Text + Puffer
        width = max(len(header) + 4, 14)
        ws.column_dimensions[get_column_letter(i)].width = width


def add_example_row_border(ws, row, num_cols):
    """Beispielzeilen mit Rahmen versehen"""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = thin_border


# ===== Workbook erstellen =====
wb = openpyxl.Workbook()

# ============================================================
# BLATT 1: Partner
# ============================================================
ws_partner = wb.active
ws_partner.title = "Partner"

partner_headers = [
    "Partner-ID",           # A
    "Vorname",              # B
    "Nachname",             # C
    "E-Mail",               # D
    "Telefon",              # E
    "Empfehlungscode",      # F
    "Status",               # G
    "Registrierung am",     # H
    "Double-Opt-In am",     # I
    "DSE-Version",          # J
    "TB-Version",           # K
    "Consent-IP",           # L
    "Anzahl Empfehlungen",  # M
    "Davon abgeschlossen",  # N
    "Offene Provision (€)", # O
    "Ausgezahlte Prov. (€)",# P
    "Gesamt-Provision (€)", # Q
    "IBAN",                 # R
    "PayPal",               # S
    "Steuernr./USt-ID",     # T
    "Notizen",              # U
]

# Header schreiben
for i, header in enumerate(partner_headers, 1):
    ws_partner.cell(row=1, column=i, value=header)

style_header_row(ws_partner, len(partner_headers))
auto_column_width(ws_partner, partner_headers)

# Spaltenbreiten manuell anpassen für bestimmte Spalten
ws_partner.column_dimensions["A"].width = 18
ws_partner.column_dimensions["D"].width = 28
ws_partner.column_dimensions["F"].width = 18
ws_partner.column_dimensions["H"].width = 18
ws_partner.column_dimensions["I"].width = 18
ws_partner.column_dimensions["U"].width = 30

# Dropdown: Status (Spalte G)
status_validation = DataValidation(
    type="list",
    formula1='"Aktiv,Pausiert,Gekündigt"',
    allow_blank=False,
)
status_validation.error = "Bitte wähle: Aktiv, Pausiert oder Gekündigt"
status_validation.errorTitle = "Ungültiger Status"
ws_partner.add_data_validation(status_validation)
status_validation.add("G2:G1000")

# Bedingte Formatierung: Status-Spalte (G)
ws_partner.conditional_formatting.add(
    "G2:G1000",
    CellIsRule(operator="equal", formula=['"Aktiv"'], fill=fill_gruen, font=font_weiss)
)
ws_partner.conditional_formatting.add(
    "G2:G1000",
    CellIsRule(operator="equal", formula=['"Pausiert"'], fill=fill_gelb, font=Font(bold=True))
)
ws_partner.conditional_formatting.add(
    "G2:G1000",
    CellIsRule(operator="equal", formula=['"Gekündigt"'], fill=fill_rot, font=font_weiss)
)

# Beispielzeile mit Formeln (Zeile 2)
ws_partner.cell(row=2, column=1, value="DW24-ANNA01")       # A: Partner-ID
ws_partner.cell(row=2, column=2, value="Anna")              # B: Vorname
ws_partner.cell(row=2, column=3, value="Müller")            # C: Nachname
ws_partner.cell(row=2, column=4, value="anna@example.com")  # D: E-Mail
ws_partner.cell(row=2, column=5, value="+49 170 1234567")   # E: Telefon
ws_partner.cell(row=2, column=6, value="DW24-ANNA01")       # F: Empfehlungscode
ws_partner.cell(row=2, column=7, value="Aktiv")             # G: Status
ws_partner.cell(row=2, column=8, value="03.03.2026")        # H: Registrierung
ws_partner.cell(row=2, column=9, value="03.03.2026")        # I: DOI
ws_partner.cell(row=2, column=10, value="v1.0")             # J: DSE
ws_partner.cell(row=2, column=11, value="v1.0")             # K: TB
ws_partner.cell(row=2, column=12, value="192.168.1.1")      # L: IP

# Formeln für berechnete Felder (ab Zeile 2)
# M: Anzahl Empfehlungen = COUNTIF aus Empfehlungen-Blatt
ws_partner.cell(row=2, column=13).value = '=COUNTIF(Empfehlungen!B:B,F2)'
# N: Davon abgeschlossen = COUNTIFS
ws_partner.cell(row=2, column=14).value = '=COUNTIFS(Empfehlungen!B:B,F2,Empfehlungen!J:J,"Abgeschlossen")'
# O: Offene Provision = (Abgeschlossen mit "Ja" bei Provision fällig, aber "Nein" bei ausgezahlt) × 199
ws_partner.cell(row=2, column=15).value = '=COUNTIFS(Empfehlungen!B:B,F2,Empfehlungen!M:M,"Ja",Empfehlungen!N:N,"Nein")*199'
ws_partner.cell(row=2, column=15).number_format = '#,##0.00 €'
# P: Ausgezahlte Provision = SUMMENPRODUKT aus Auszahlungen
ws_partner.cell(row=2, column=16).value = '=SUMPRODUCT((Auszahlungen!B:B=F2)*Auszahlungen!E:E)'
ws_partner.cell(row=2, column=16).number_format = '#,##0.00 €'
# Q: Gesamt-Provision = O + P
ws_partner.cell(row=2, column=17).value = '=O2+P2'
ws_partner.cell(row=2, column=17).number_format = '#,##0.00 €'

add_example_row_border(ws_partner, 2, len(partner_headers))

# Sensible Spalten ausblenden (R, S, T)
ws_partner.column_dimensions["R"].hidden = True
ws_partner.column_dimensions["S"].hidden = True
ws_partner.column_dimensions["T"].hidden = True


# ============================================================
# BLATT 2: Empfehlungen
# ============================================================
ws_empf = wb.create_sheet("Empfehlungen")

empf_headers = [
    "Empfehlungs-ID",       # A
    "Empfehlungscode",      # B
    "Partner-Name",         # C
    "Empfohlener Name",     # D
    "Empfohlene Firma",     # E
    "Empfohlene E-Mail",    # F
    "Empfohlenes Telefon",  # G
    "Branche/Gewerk",       # H
    "Empfohlen am",         # I
    "Status",               # J
    "DW24-Auftragswert (€)",# K
    "Kunde bezahlt am",     # L
    "Provision fällig",     # M
    "Provision ausgezahlt", # N
    "Notizen",              # O
]

for i, header in enumerate(empf_headers, 1):
    ws_empf.cell(row=1, column=i, value=header)

style_header_row(ws_empf, len(empf_headers))
auto_column_width(ws_empf, empf_headers)

ws_empf.column_dimensions["A"].width = 18
ws_empf.column_dimensions["D"].width = 22
ws_empf.column_dimensions["E"].width = 22
ws_empf.column_dimensions["F"].width = 26
ws_empf.column_dimensions["K"].width = 22
ws_empf.column_dimensions["O"].width = 30

# Dropdown: Status (Spalte J)
empf_status_validation = DataValidation(
    type="list",
    formula1='"Neu,Kontaktiert,Angebot gesendet,Abgeschlossen,Abgelehnt"',
    allow_blank=False,
)
ws_empf.add_data_validation(empf_status_validation)
empf_status_validation.add("J2:J1000")

# Dropdown: Provision ausgezahlt (Spalte N)
ja_nein_validation = DataValidation(
    type="list",
    formula1='"Ja,Nein"',
    allow_blank=True,
)
ws_empf.add_data_validation(ja_nein_validation)
ja_nein_validation.add("N2:N1000")

# Bedingte Formatierung: Status (Spalte J)
ws_empf.conditional_formatting.add(
    "J2:J1000",
    CellIsRule(operator="equal", formula=['"Neu"'], fill=fill_blau, font=font_weiss)
)
ws_empf.conditional_formatting.add(
    "J2:J1000",
    CellIsRule(operator="equal", formula=['"Kontaktiert"'], fill=fill_gelb, font=Font(bold=True))
)
ws_empf.conditional_formatting.add(
    "J2:J1000",
    CellIsRule(operator="equal", formula=['"Angebot gesendet"'], fill=fill_orange_hell, font=Font(bold=True))
)
ws_empf.conditional_formatting.add(
    "J2:J1000",
    CellIsRule(operator="equal", formula=['"Abgeschlossen"'], fill=fill_gruen, font=font_weiss)
)
ws_empf.conditional_formatting.add(
    "J2:J1000",
    CellIsRule(operator="equal", formula=['"Abgelehnt"'], fill=fill_rot, font=font_weiss)
)

# Beispielzeile
ws_empf.cell(row=2, column=1, value="EMP-2026-001")
ws_empf.cell(row=2, column=2, value="DW24-ANNA01")
# C: Partner-Name per VLOOKUP
ws_empf.cell(row=2, column=3).value = '=IFERROR(VLOOKUP(B2,Partner!F:C,1,FALSE)&" "&VLOOKUP(B2,Partner!F:B,1,FALSE),"")'
# Vereinfachter VLOOKUP: INDEX/MATCH
ws_empf.cell(row=2, column=3).value = '=IFERROR(INDEX(Partner!B:B,MATCH(B2,Partner!F:F,0))&" "&INDEX(Partner!C:C,MATCH(B2,Partner!F:F,0)),"")'
ws_empf.cell(row=2, column=4, value="Max Maier")
ws_empf.cell(row=2, column=5, value="Maier Elektrotechnik")
ws_empf.cell(row=2, column=6, value="max@maier-elektro.de")
ws_empf.cell(row=2, column=7, value="+49 171 9876543")
ws_empf.cell(row=2, column=8, value="Elektriker")
ws_empf.cell(row=2, column=9, value="05.03.2026")
ws_empf.cell(row=2, column=10, value="Abgeschlossen")
ws_empf.cell(row=2, column=11, value=1490)
ws_empf.cell(row=2, column=11).number_format = '#,##0.00 €'
ws_empf.cell(row=2, column=12, value="15.03.2026")
# M: Provision fällig = Wenn Kunde bezahlt hat
ws_empf.cell(row=2, column=13).value = '=IF(L2<>"","Ja","Nein")'
ws_empf.cell(row=2, column=14, value="Nein")

add_example_row_border(ws_empf, 2, len(empf_headers))


# ============================================================
# BLATT 3: Auszahlungen
# ============================================================
ws_ausz = wb.create_sheet("Auszahlungen")

ausz_headers = [
    "Auszahlungs-ID",      # A
    "Empfehlungscode",      # B
    "Partner-Name",         # C
    "Empfehlungs-ID",       # D
    "Betrag (€)",           # E
    "Auszahlungsdatum",     # F
    "Zahlungsmethode",      # G
    "Referenz/Buchungsnr.", # H
    "Status",               # I
    "Notizen",              # J
]

for i, header in enumerate(ausz_headers, 1):
    ws_ausz.cell(row=1, column=i, value=header)

style_header_row(ws_ausz, len(ausz_headers))
auto_column_width(ws_ausz, ausz_headers)

ws_ausz.column_dimensions["A"].width = 18
ws_ausz.column_dimensions["C"].width = 20
ws_ausz.column_dimensions["D"].width = 18
ws_ausz.column_dimensions["H"].width = 22
ws_ausz.column_dimensions["J"].width = 30

# Dropdown: Zahlungsmethode (Spalte G)
zahlungs_validation = DataValidation(
    type="list",
    formula1='"Überweisung,PayPal"',
    allow_blank=True,
)
ws_ausz.add_data_validation(zahlungs_validation)
zahlungs_validation.add("G2:G1000")

# Dropdown: Status (Spalte I)
ausz_status_validation = DataValidation(
    type="list",
    formula1='"Offen,Angewiesen,Bestätigt"',
    allow_blank=False,
)
ws_ausz.add_data_validation(ausz_status_validation)
ausz_status_validation.add("I2:I1000")

# Bedingte Formatierung: Auszahlungs-Status
ws_ausz.conditional_formatting.add(
    "I2:I1000",
    CellIsRule(operator="equal", formula=['"Offen"'], fill=fill_gelb, font=Font(bold=True))
)
ws_ausz.conditional_formatting.add(
    "I2:I1000",
    CellIsRule(operator="equal", formula=['"Angewiesen"'], fill=fill_blau, font=font_weiss)
)
ws_ausz.conditional_formatting.add(
    "I2:I1000",
    CellIsRule(operator="equal", formula=['"Bestätigt"'], fill=fill_gruen, font=font_weiss)
)

# Beispielzeile
ws_ausz.cell(row=2, column=1, value="AUS-2026-001")
ws_ausz.cell(row=2, column=2, value="DW24-ANNA01")
# C: Partner-Name per INDEX/MATCH
ws_ausz.cell(row=2, column=3).value = '=IFERROR(INDEX(Partner!B:B,MATCH(B2,Partner!F:F,0))&" "&INDEX(Partner!C:C,MATCH(B2,Partner!F:F,0)),"")'
ws_ausz.cell(row=2, column=4, value="EMP-2026-001")
ws_ausz.cell(row=2, column=5, value=199)
ws_ausz.cell(row=2, column=5).number_format = '#,##0.00 €'
ws_ausz.cell(row=2, column=6, value="20.03.2026")
ws_ausz.cell(row=2, column=7, value="Überweisung")
ws_ausz.cell(row=2, column=8, value="")
ws_ausz.cell(row=2, column=9, value="Offen")

add_example_row_border(ws_ausz, 2, len(ausz_headers))


# ============================================================
# BLATT 4: Dashboard
# ============================================================
ws_dash = wb.create_sheet("Dashboard")

# Dashboard-Layout: Große Kennzahlen
title_font = Font(name="Calibri", bold=True, size=24, color=DW24_DUNKEL)
label_font = Font(name="Calibri", size=12, color="64748B")
value_font = Font(name="Calibri", bold=True, size=28, color=DW24_ORANGE)
section_font = Font(name="Calibri", bold=True, size=16, color=DW24_DUNKEL)

# Titel
ws_dash.cell(row=1, column=1, value="📊 DW24 Empfehlungsprogramm – Dashboard")
ws_dash.cell(row=1, column=1).font = title_font
ws_dash.merge_cells("A1:F1")

# Spaltenbreiten
for col in ["A", "B", "C", "D", "E", "F"]:
    ws_dash.column_dimensions[col].width = 22

ws_dash.row_dimensions[1].height = 40
ws_dash.row_dimensions[3].height = 20
ws_dash.row_dimensions[4].height = 45
ws_dash.row_dimensions[6].height = 20
ws_dash.row_dimensions[7].height = 45
ws_dash.row_dimensions[9].height = 20
ws_dash.row_dimensions[10].height = 45

# ---- Zeile 3-4: Hauptkennzahlen ----
kpis_row1 = [
    ("Aktive Partner", '=COUNTIF(Partner!G:G,"Aktiv")'),
    ("Empfehlungen gesamt", '=COUNTA(Empfehlungen!A:A)-1'),
    ("Davon abgeschlossen", '=COUNTIF(Empfehlungen!J:J,"Abgeschlossen")'),
]

for i, (label, formula) in enumerate(kpis_row1):
    col = (i * 2) + 1  # Spalten A, C, E
    ws_dash.cell(row=3, column=col, value=label).font = label_font
    ws_dash.cell(row=4, column=col).value = formula
    ws_dash.cell(row=4, column=col).font = value_font
    ws_dash.cell(row=4, column=col).alignment = Alignment(horizontal="left")

# ---- Zeile 6-7: Finanzkennzahlen ----
kpis_row2 = [
    ("Conversion-Rate", '=IFERROR(COUNTIF(Empfehlungen!J:J,"Abgeschlossen")/(COUNTA(Empfehlungen!A:A)-1),0)'),
    ("Offene Provisionen (€)", '=COUNTIFS(Empfehlungen!M:M,"Ja",Empfehlungen!N:N,"Nein")*199'),
    ("Ausgezahlt gesamt (€)", "=SUM(Auszahlungen!E:E)"),
]

for i, (label, formula) in enumerate(kpis_row2):
    col = (i * 2) + 1
    ws_dash.cell(row=6, column=col, value=label).font = label_font
    ws_dash.cell(row=7, column=col).value = formula
    ws_dash.cell(row=7, column=col).font = value_font
    ws_dash.cell(row=7, column=col).alignment = Alignment(horizontal="left")

# Conversion-Rate als Prozent formatieren
ws_dash.cell(row=7, column=1).number_format = '0.0%'
# Währungsformat
ws_dash.cell(row=7, column=3).number_format = '#,##0.00 €'
ws_dash.cell(row=7, column=5).number_format = '#,##0.00 €'

# ---- Zeile 9-10: Monatskennzahlen ----
kpis_row3 = [
    ("Empfehlungen diesen Monat", '=COUNTIFS(Empfehlungen!I:I,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Empfehlungen!I:I,"<="&EOMONTH(TODAY(),0))'),
    ("Neue Partner diesen Monat", '=COUNTIFS(Partner!H:H,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Partner!H:H,"<="&EOMONTH(TODAY(),0))'),
    ("Umsatz diesen Monat (€)", '=SUMPRODUCT((Empfehlungen!L:L>=DATE(YEAR(TODAY()),MONTH(TODAY()),1))*(Empfehlungen!L:L<=EOMONTH(TODAY(),0))*Empfehlungen!K:K)'),
]

for i, (label, formula) in enumerate(kpis_row3):
    col = (i * 2) + 1
    ws_dash.cell(row=9, column=col, value=label).font = label_font
    ws_dash.cell(row=10, column=col).value = formula
    ws_dash.cell(row=10, column=col).font = value_font
    ws_dash.cell(row=10, column=col).alignment = Alignment(horizontal="left")

ws_dash.cell(row=10, column=5).number_format = '#,##0.00 €'

# ---- Trennlinie ----
ws_dash.cell(row=12, column=1, value="🏆 Top-5 Partner nach Empfehlungen")
ws_dash.cell(row=12, column=1).font = section_font
ws_dash.merge_cells("A12:D12")

# Top-5 Header
top5_headers = ["Rang", "Partner-Name", "Empfehlungscode", "Empfehlungen"]
for i, h in enumerate(top5_headers, 1):
    cell = ws_dash.cell(row=13, column=i, value=h)
    cell.font = Font(bold=True, size=11, color=DW24_WEISS)
    cell.fill = PatternFill(start_color=DW24_DUNKEL, end_color=DW24_DUNKEL, fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Top-5 Formeln (mit LARGE/INDEX)
for rank in range(1, 6):
    row = 13 + rank
    ws_dash.cell(row=row, column=1, value=rank)
    ws_dash.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=row, column=1).font = Font(bold=True, size=14, color=DW24_ORANGE)

    # Partner-Name (INDEX/MATCH auf Basis LARGE)
    ws_dash.cell(row=row, column=2).value = f'=IFERROR(INDEX(Partner!B:B,MATCH(LARGE(Partner!M:M,{rank}),Partner!M:M,0))&" "&INDEX(Partner!C:C,MATCH(LARGE(Partner!M:M,{rank}),Partner!M:M,0)),"-")'
    ws_dash.cell(row=row, column=3).value = f'=IFERROR(INDEX(Partner!F:F,MATCH(LARGE(Partner!M:M,{rank}),Partner!M:M,0)),"-")'
    ws_dash.cell(row=row, column=4).value = f'=IFERROR(LARGE(Partner!M:M,{rank}),"-")'
    ws_dash.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=row, column=4).font = Font(bold=True, size=12)

    for col in range(1, 5):
        ws_dash.cell(row=row, column=col).border = thin_border

# ---- Hinweistext ----
ws_dash.cell(row=20, column=1, value="ℹ️ Alle Werte werden automatisch berechnet. Keine manuellen Eingaben auf diesem Blatt.")
ws_dash.cell(row=20, column=1).font = Font(italic=True, color="94A3B8", size=10)
ws_dash.merge_cells("A20:F20")

# Blattschutz für Dashboard (nur Lesen)
ws_dash.protection.sheet = True
ws_dash.protection.password = "dw24dash"
ws_dash.protection.enable()


# ============================================================
# Speichern
# ============================================================
output_path = "/Users/manuelhorn/Documents/Kunden/Digitalwerk24-Website/Digitalwerk24_Vertriebspartner-Seite/DW24-Empfehlungsprogramm.xlsx"
wb.save(output_path)
print(f"✅ XLSX erstellt: {output_path}")
print(f"   📋 4 Blätter: Partner, Empfehlungen, Auszahlungen, Dashboard")
print(f"   📊 Formeln, Dropdowns, bedingte Formatierung enthalten")
print(f"   🔒 Dashboard ist schreibgeschützt (Passwort: dw24dash)")
print(f"   🙈 Sensible Spalten (IBAN, PayPal, Steuernr.) ausgeblendet")
print(f"\n   → In Google Drive hochladen und als Google Sheets öffnen")
